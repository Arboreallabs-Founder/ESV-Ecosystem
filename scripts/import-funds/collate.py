"""
Collate the ESV fund workbooks into one normalised, import-ready dataset.

Sources (all three are needed — none is a superset of the others):
  ESV Database.xlsx / Data          264 rows, the master
  ESV Database.xlsx / Funds - KR     69 rows, 29 of which exist nowhere else
  ESV Database.xlsx / Funds - AR     81 rows, 3 unique + the only Connect Strength data
  Fund Completeness Check.xlsx      264 rows, the POC employment audit

Writes funds.json (the data) and report.md (what a human needs to look at). Touches no database:
run collate, read the report, then run load.py.

Usage:  python scripts/import-funds/collate.py
"""
import json
import os
import re
from collections import Counter, defaultdict

import openpyxl

BASE = r"C:\Users\upadh\Documents\Earlyseed Ventures\Data"
OUT = os.path.dirname(os.path.abspath(__file__))

# ── ESV people ───────────────────────────────────────────────────────────────
# Initials as used in the sheets. `user` is the display name of a current user record; None means
# they have left and have no account, so attribution is kept as a name rather than a foreign key.
# Losing it is not an option — NB alone owns 152 of the funds.
ESV_PEOPLE = {
    'NB': ('Neeti Bokaria',      False),
    'SB': ('Siddhant Baliga',    True),
    'KS': ('Karan Shah',         True),
    'KR': ('Kalpak Roy',         False),
    'MG': ('Monica Gupta',       True),
    'MP': ('Manan Patel',        False),
    'AR': ('Arjun Renapurkar',   False),
    'RG': ('Ruhaan Gupta',       True),
}

# ── Service type ─────────────────────────────────────────────────────────────
# Maps the 73 free-text values onto the 12 the app's CHECK constraint allows.
SERVICE_TYPE_MAP = {
    'fund': 'vc_fund', 'venture fund': 'vc_fund', 'vc': 'vc_fund', 'vc fund': 'vc_fund',
    'venture capital': 'vc_fund', 'micro vc': 'vc_fund', 'micro fund': 'vc_fund',
    # An AIF Category II in India is the VC/PE vehicle; it is a regulatory class, not a strategy.
    'cat 2 fund': 'vc_fund', 'cat ii fund': 'vc_fund', 'category 2 fund': 'vc_fund',
    'angel fund': 'angel_fund', 'angel network': 'angel_fund', 'angel syndicate': 'angel_fund',
    'angel investor': 'angel_investor', 'angel': 'angel_investor',
    'family office': 'family_office', 'family fund': 'family_office',
    'pe fund': 'private_equity', 'private equity': 'private_equity', 'pe': 'private_equity',
    'growth fund': 'growth_equity', 'growth equity': 'growth_equity',
    'debt fund': 'debt_fund', 'venture debt': 'debt_fund', 'debt': 'debt_fund',
    'accelerator': 'accelerator', 'incubator': 'accelerator', 'accelerator/incubator': 'accelerator',
    'fund of funds': 'fund_of_funds', 'fof': 'fund_of_funds',
    'corporate vc': 'corporate_vc', 'cvc': 'corporate_vc', 'corporate fund': 'corporate_vc',
    'sovereign wealth': 'sovereign_wealth',
    'investment bank': 'merchant_bank', 'investment banking': 'merchant_bank',
    'merchant bank': 'merchant_bank', 'ib': 'merchant_bank',
}

# ── Stage ladder ─────────────────────────────────────────────────────────────
STAGE_LADDER = ['pre_seed', 'seed', 'pre_series_a', 'series_a', 'series_b', 'growth']
STAGE_ALIASES = [
    (r'pre[\s-]*series[\s-]*a', 'pre_series_a'),
    (r'pre[\s-]*seed',          'pre_seed'),
    (r'series[\s-]*b\+?',       'series_b'),
    (r'series[\s-]*a\+?',       'series_a'),
    (r'\bgrowth\b',             'growth'),
    (r'\blate stage\b',         'growth'),
    (r'\bseed\b',               'seed'),
    (r'\bearly\b',              'seed'),
]

def parse_stage(raw):
    """Free text -> (min, max) on the ladder. 106 distinct spellings collapse to a range."""
    if not raw:
        return None, None, None
    s = str(raw).strip().lower()
    if s in ('-', 'na', 'n/a', ''):
        return None, None, None
    if 'agnostic' in s or 'all stage' in s:
        return 'pre_seed', 'growth', None
    found = []
    for pat, val in STAGE_ALIASES:
        for m in re.finditer(pat, s):
            found.append((m.start(), val))
    if not found:
        return None, None, str(raw).strip()   # keep the original for review
    order = {v: i for i, v in enumerate(STAGE_LADDER)}
    vals = sorted({v for _, v in found}, key=lambda v: order[v])
    return vals[0], vals[-1], None

# ── Ticket size ──────────────────────────────────────────────────────────────
CRORE, LAKH = 10_000_000, 100_000

def parse_ticket(raw):
    """
    -> (min, max, currency, unparsed_note)

    Currency is only set when the cell says so. 92 of 245 cells give a bare range like
    "1,000,000 to 5,000,000" — guessing there is an 80x error that would silently poison every
    ticket-size filter, so those import as NULL with the original text kept for review.
    """
    if raw is None:
        return None, None, None, None
    s = str(raw).strip()
    if s in ('-', '', 'NA', 'N/A', 'TBD'):
        return None, None, None, None
    low = s.lower()

    currency = None
    if '$' in s or 'usd' in low or 'mn usd' in low:
        currency = 'USD'
    elif 'inr' in low or '₹' in s or 'rs' in low or ' cr' in low or 'crore' in low or 'lakh' in low or 'lacs' in low:
        currency = 'INR'

    nums = []
    for m in re.finditer(r'([\d][\d,\.]*)\s*(crs?|crore s?|cr|lakhs?|lacs?|mn|m|k)?', low):
        txt = m.group(1).replace(',', '')
        if not txt or txt == '.':
            continue
        try:
            val = float(txt)
        except ValueError:
            continue
        unit = (m.group(2) or '').strip()
        if unit in ('cr', 'crs', 'crore', 'crore s', 'crores'):
            val *= CRORE
        elif unit in ('lakh', 'lakhs', 'lac', 'lacs'):
            val *= LAKH
        elif unit in ('mn', 'm'):
            val *= 1_000_000
        elif unit == 'k':
            val *= 1_000
        if val > 0:
            nums.append(val)

    if not nums:
        return None, None, currency, s
    lo, hi = min(nums), max(nums)
    if currency is None:
        # Parsed fine but we do not know the unit — per the decision, import blank and keep the
        # text so someone can review it.
        return None, None, None, s
    return lo, (hi if hi != lo else None), currency, None

# ── Sectors ──────────────────────────────────────────────────────────────────
CANON = {
    'healthtech': 'HealthTech', 'health tech': 'HealthTech', 'healthcare': 'HealthTech',
    'health care': 'HealthTech', 'health': 'HealthTech', 'medtech': 'HealthTech',
    'fintech': 'FinTech', 'fin tech': 'FinTech', 'financial services': 'FinTech',
    'insuretech': 'FinTech', 'insurtech': 'FinTech',
    'saas': 'SaaS', 'b2b saas': 'SaaS', 'enterprise saas': 'SaaS', 'software': 'SaaS',
    'deeptech': 'DeepTech', 'deep tech': 'DeepTech', 'deep science': 'DeepTech',
    'frontier tech': 'DeepTech', 'space tech': 'SpaceTech', 'spacetech': 'SpaceTech',
    'ai': 'AI/ML', 'ml': 'AI/ML', 'ai/ml': 'AI/ML', 'artificial intelligence': 'AI/ML',
    'genai': 'AI/ML', 'gen ai': 'AI/ML',
    'agritech': 'AgriTech', 'agri tech': 'AgriTech', 'agriculture': 'AgriTech', 'agri': 'AgriTech',
    'foodtech': 'FoodTech', 'food tech': 'FoodTech', 'food': 'FoodTech', 'f&b': 'FoodTech',
    'edtech': 'EdTech', 'ed tech': 'EdTech', 'education': 'EdTech',
    'climatetech': 'ClimateTech', 'climate tech': 'ClimateTech', 'climate': 'ClimateTech',
    'cleantech': 'ClimateTech', 'clean tech': 'ClimateTech', 'sustainability': 'ClimateTech',
    'esg': 'ClimateTech', 'renewable energy': 'ClimateTech', 'energy': 'Energy',
    'ev': 'EV & Mobility', 'mobility': 'EV & Mobility', 'automotive': 'EV & Mobility',
    'logistics': 'Logistics', 'logistics tech': 'Logistics', 'supply chain': 'Logistics',
    'gaming': 'Gaming', 'esports': 'Gaming',
    'consumer': 'Consumer', 'consumer tech': 'Consumer', 'consumertech': 'Consumer',
    'consumer brands': 'Consumer', 'd2c': 'D2C', 'b2c': 'D2C', 'b2c commerce': 'D2C',
    'e-commerce': 'E-commerce', 'ecommerce': 'E-commerce', 'commerce': 'E-commerce',
    'marketplace': 'Marketplace', 'marketplaces': 'Marketplace',
    'b2b': 'B2B', 'b2b tech': 'B2B', 'b2btech': 'B2B', 'enterprise': 'B2B',
    'web3': 'Web3', 'blockchain': 'Web3', 'crypto': 'Web3',
    'media': 'Media', 'entertainment': 'Media', 'content': 'Media',
    'manufacturing': 'Manufacturing', 'industrial': 'Manufacturing', 'robotics': 'Robotics',
    'biotech': 'BioTech', 'life sciences': 'BioTech', 'pharma': 'BioTech',
    'real estate': 'Real Estate', 'proptech': 'Real Estate', 'construction': 'Real Estate',
    'defence': 'Defence', 'defense': 'Defence', 'semiconductor': 'Semiconductors',
    'hrtech': 'HRTech', 'hr tech': 'HRTech', 'traveltech': 'Travel', 'travel': 'Travel',
    'retail': 'Retail', 'fashion': 'Fashion', 'beauty': 'Beauty', 'sports': 'Sports',
    'cybersecurity': 'Cybersecurity', 'security': 'Cybersecurity',
    'iot': 'IoT', 'hardware': 'Hardware', 'tech': None, 'technology': None,   # too vague to keep
    'agnostic': 'Agnostic', 'agnsotic': 'Agnostic',
    # Second pass, from the words the first run reported as unmapped.
    'edutech': 'EdTech', 'finance': 'FinTech', 'regtech': 'FinTech', 'insurance': 'FinTech',
    'cybersec': 'Cybersecurity', 'infra': 'Infrastructure', 'infrastructure': 'Infrastructure',
    'brands': 'Consumer', 'lifestyle': 'Consumer', 'personal care': 'Consumer',
    'clothing': 'Fashion', 'home': 'Consumer', 'fnb': 'FoodTech',
    'sportstech': 'Sports', 'fitness': 'HealthTech',
    'drone tech': 'Drones', 'drones': 'Drones', 'aerospace': 'SpaceTech',
    'transport': 'EV & Mobility', 'transportation': 'EV & Mobility',
    'warehousing': 'Logistics', 'logistic': 'Logistics',
    'real-estate': 'Real Estate', 'propertytech': 'Real Estate', 'prop tech': 'Real Estate',
    'creator-economy': 'Media', 'creator economy': 'Media', 'social': 'Media',
    'legal': 'LegalTech', 'legaltech': 'LegalTech',
    'ar': 'AR/VR', 'vr': 'AR/VR', 'ar/vr': 'AR/VR',
    'dev-tools': 'SaaS', 'dev tools': 'SaaS', 'dev infra': 'SaaS',
    'b2b2c': 'B2B', 'enterprisetech': 'B2B', 'enterprise tech': 'B2B',
    'hospitality': 'Travel', 'environment': 'ClimateTech', 'waste management': 'ClimateTech',
    'quantum computing': 'DeepTech', 'synthetic biology': 'BioTech',
    'autonomous vehicles': 'EV & Mobility', 'cleanenergy': 'Energy',
}
# Words that mark everything after them as an exclusion, not a preference.
#
# "no" must NOT match inside "no code" / "no-code" — that is a technology category, and reading it
# as a negation inverted an entire fund's sector list (Speciale Invest) into exclusions on the
# first run. The lookahead is the whole reason this is not a plain word list.
NEG_MARKERS = re.compile(
    r'\b(no(?!\s*[- ]?code)|not|avoid|except|excluding|exclude|apart from|other than|steer clear of)\b',
    re.I,
)

# Things funds exclude that are not investment sectors, so they never appear in CANON. Without
# these the exclusion list either loses the real signal ("no meat") or fills with noise.
EXCLUSION_EXTRA = {
    'meat': 'Meat', 'alcohol': 'Alcohol', 'alchohol': 'Alcohol', 'alc': 'Alcohol',
    'liquor': 'Alcohol', 'tobacco': 'Tobacco', 'gambling': 'Gambling', 'betting': 'Gambling',
    'real money gaming': 'Real-money gaming', 'rmg': 'Real-money gaming',
    'adult': 'Adult content', 'weapons': 'Weapons', 'arms': 'Weapons',
    'hardware heavy': 'Hardware-heavy', 'hardware heavy cos': 'Hardware-heavy',
    'consumer brands': 'Consumer',
}

# Filler that survives tokenising but carries no meaning.
STOPWORDS = re.compile(r'\b(etc|and|or|the|they|we|is|are|in|of|for|to|a|an|very|active|interested)\b', re.I)


def _map_token(t, negative):
    """
    A token -> a canonical tag, or None if it is not one.

    Exclusions are held to the SAME standard as inclusions: only recognised vocabulary gets in.
    An exclusion list containing "they" or "very active" is worse than no list at all, because it
    looks authoritative while filtering on nonsense.
    """
    t = STOPWORDS.sub(' ', str(t).strip().strip('.-').lower())
    t = re.sub(r'\s+', ' ', t).strip()
    if not t or len(t) > 40 or t.isdigit():
        return None
    if t in CANON:
        return CANON[t]
    if negative and t in EXCLUSION_EXTRA:
        return EXCLUSION_EXTRA[t]
    # "edtechs", "marketplaces", "brands" — try the singular before giving up.
    if t.endswith('s') and t[:-1] in CANON:
        return CANON[t[:-1]]
    if negative and t.endswith('s') and t[:-1] in EXCLUSION_EXTRA:
        return EXCLUSION_EXTRA[t[:-1]]
    # "interested in deeptech", once the marker is stripped, still means DeepTech. Look for a known
    # tag inside the phrase rather than discarding the whole token.
    vocabs = [CANON, EXCLUSION_EXTRA] if negative else [CANON]
    for vocab in vocabs:
        for k, v in vocab.items():
            if v and len(k) > 3 and re.search(rf'\b{re.escape(k)}\b', t):
                return v
    return None


def parse_sectors(raw):
    """
    -> (included, excluded, leftovers)

    The cells mix both directions: "Agnostic - Healthtech, SAAS - NO Marketplace, meat, alc,
    gambling". Reading only the positives would put a meat startup in front of the fund that
    wrote "no meat", which is the exact mistake this feature is meant to prevent.
    """
    if not raw:
        return [], [], []
    text = str(raw).strip()
    inc, exc, left = [], [], []

    # Split into segments and decide the polarity of each; a negative marker flips everything
    # after it within that segment.
    segments = re.split(r'\s+-\s+|\s*[;|]\s*|\.\s+', text)
    for seg in segments:
        neg = bool(NEG_MARKERS.search(seg))
        body = NEG_MARKERS.sub(' ', seg) if neg else seg
        for tok in re.split(r'[,/&]', body):
            mapped = _map_token(tok, neg)
            if mapped:
                (exc if neg else inc).append(mapped)
            elif not neg:
                t = tok.strip().strip('.-').strip().lower()
                if t and 1 < len(t) <= 40 and not t.isdigit():
                    left.append(t)

    dedupe = lambda xs: list(dict.fromkeys(xs))
    return dedupe(inc), dedupe(exc), dedupe(left)


def norm_name(s):
    return re.sub(r'[^a-z0-9]', '', str(s or '').lower())


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ('', '-', 'NA', 'N/A', 'na', '#REF!'):
        return None
    return s


def split_people(cell):
    """'Sparsh Sehgal + Kushal Bhagia' / 'Bhairavi Nagda /\\nGaurav Thakkar' -> two names."""
    if not cell:
        return []
    return [p.strip() for p in re.split(r'\s*[/+&]\s*|\s*\n\s*', str(cell)) if p.strip() and p.strip() != '-']


def split_statuses(cell):
    """'No / Yes' -> ['no','yes'], matched to people by position."""
    if not cell:
        return []
    return [p.strip().lower() for p in re.split(r'\s*[/+]\s*|\s*\n\s*', str(cell)) if p.strip()]


def status_to_employment(s):
    """
    Per the decision: 'not found' counts as moved on. The audit wording is preserved on the record
    so a not-found is still distinguishable from an observed departure.
    """
    if not s:
        return 'unknown', None
    s = s.strip().lower()
    if s.startswith('y'):
        return 'active', None
    if s.startswith('n') and 'not found' not in s:
        return 'moved_on', None
    if 'not found' in s:
        return 'moved_on', 'Not found during the audit — treated as moved on.'
    return 'unknown', None


def main():
    wb = openpyxl.load_workbook(os.path.join(BASE, 'ESV Database.xlsx'), data_only=True)
    funds = {}          # norm name -> record
    report = defaultdict(list)

    # Placeholder rows left behind in the sheets — not funds.
    JUNK = {'test', 'tests', 'testing', 'na', 'tbd', 'xxx', 'sample'}

    def get_or_create(name, source):
        key = norm_name(name)
        if key not in funds:
            funds[key] = {
                'name': str(name).strip(), 'sources': [], 'country': None, 'website': None,
                'sectors_raw': None, 'esv_pocs': [], 'service_type_raw': None,
                'ticket_raw': None, 'stage_raw': None, 'notes': [], 'connect_strength': None,
                'contacts': [],
            }
        if source not in funds[key]['sources']:
            funds[key]['sources'].append(source)
        return funds[key]

    # ── Master: Data ──
    ws = wb['Data']
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(r, 1).value)
        if not name or name.strip().lower() in JUNK:
            continue
        f = get_or_create(name, 'Data')
        f['country'] = f['country'] or clean(ws.cell(r, 2).value)
        f['website'] = f['website'] or clean(ws.cell(r, 3).value)
        f['sectors_raw'] = f['sectors_raw'] or clean(ws.cell(r, 4).value)
        f['service_type_raw'] = f['service_type_raw'] or clean(ws.cell(r, 6).value)
        f['ticket_raw'] = f['ticket_raw'] or clean(ws.cell(r, 7).value)
        f['stage_raw'] = f['stage_raw'] or clean(ws.cell(r, 8).value)
        for init in re.split(r'[,/]', str(clean(ws.cell(r, 5).value) or '')):
            init = init.strip().upper()
            if init in ESV_PEOPLE and init not in f['esv_pocs']:
                f['esv_pocs'].append(init)
        for note_col in (21, 22):
            n = clean(ws.cell(r, note_col).value)
            if n and n not in f['notes']:
                f['notes'].append(n)
        # Two POC blocks per row.
        for base_c, order in ((9, 1), (15, 2)):
            pname = clean(ws.cell(r, base_c).value)
            if not pname:
                continue
            people = split_people(pname)
            for idx, person in enumerate(people):
                f['contacts'].append({
                    'name': person,
                    'role': clean(ws.cell(r, base_c + 1).value),
                    'linkedin_url': clean(ws.cell(r, base_c + 2).value),
                    'linkedin_status': clean(ws.cell(r, base_c + 3).value),
                    'phone': clean(ws.cell(r, base_c + 4).value),
                    'email': clean(ws.cell(r, base_c + 5).value),
                    'slot': order,
                    'multi_index': idx,
                })

    # ── Funds - KR: 29 funds that exist nowhere else ──
    ws = wb['Funds - KR']
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(r, 2).value)
        if not name:
            continue
        f = get_or_create(name, 'Funds - KR')
        f['country'] = f['country'] or clean(ws.cell(r, 3).value)
        f['website'] = f['website'] or clean(ws.cell(r, 4).value)
        f['sectors_raw'] = f['sectors_raw'] or clean(ws.cell(r, 5).value)
        f['service_type_raw'] = f['service_type_raw'] or clean(ws.cell(r, 10).value)
        for init in re.split(r'[,/]', str(clean(ws.cell(r, 6).value) or '')):
            init = init.strip().upper()
            if init in ESV_PEOPLE and init not in f['esv_pocs']:
                f['esv_pocs'].append(init)
        for c in (11, 16):
            n = clean(ws.cell(r, c).value)
            if n and n not in f['notes']:
                f['notes'].append(n)
        pname = clean(ws.cell(r, 7).value)
        if pname and not any(norm_name(x['name']) == norm_name(pname) for x in f['contacts']):
            for person in split_people(pname):
                f['contacts'].append({
                    'name': person, 'role': None,
                    'linkedin_url': clean(ws.cell(r, 12).value),
                    'linkedin_status': clean(ws.cell(r, 13).value),
                    'phone': clean(ws.cell(r, 8).value),
                    'email': clean(ws.cell(r, 9).value),
                    'slot': 1, 'multi_index': 0,
                })

    # ── Funds - AR: 3 unique, and the only Connect Strength anywhere ──
    ws = wb['Funds - AR']
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(r, 4).value)
        if not name:
            continue
        f = get_or_create(name, 'Funds - AR')
        f['connect_strength'] = (clean(ws.cell(r, 3).value) or '').lower() or None
        f['country'] = f['country'] or clean(ws.cell(r, 6).value)
        f['sectors_raw'] = f['sectors_raw'] or clean(ws.cell(r, 8).value) or clean(ws.cell(r, 7).value)
        f['ticket_raw'] = f['ticket_raw'] or clean(ws.cell(r, 9).value)
        f['stage_raw'] = f['stage_raw'] or clean(ws.cell(r, 10).value)
        for init in re.split(r'[,/]', str(clean(ws.cell(r, 2).value) or '')):
            init = init.strip().upper()
            if init in ESV_PEOPLE and init not in f['esv_pocs']:
                f['esv_pocs'].append(init)
        n = clean(ws.cell(r, 11).value)
        if n and n not in f['notes']:
            f['notes'].append(n)
        pname = clean(ws.cell(r, 12).value)
        if pname and not any(norm_name(x['name']) == norm_name(pname) for x in f['contacts']):
            for person in split_people(pname):
                f['contacts'].append({
                    'name': person, 'role': None,
                    'linkedin_url': clean(ws.cell(r, 15).value),
                    'linkedin_status': None,
                    'phone': clean(ws.cell(r, 14).value),
                    'email': clean(ws.cell(r, 13).value),
                    'slot': 1, 'multi_index': 0,
                })
    wb.close()

    # ── The employment audit ──
    wb2 = openpyxl.load_workbook(os.path.join(BASE, 'Fund Completeness Check.xlsx'), data_only=True)
    ws = wb2['Sheet1']
    audit = {}
    for r in range(2, ws.max_row + 1):
        fund = clean(ws.cell(r, 1).value)
        if not fund:
            continue
        audit.setdefault(norm_name(fund), []).append({
            'poc1': clean(ws.cell(r, 2).value),
            'designation': clean(ws.cell(r, 3).value),
            'new_company': clean(ws.cell(r, 4).value),
            'new_designation': clean(ws.cell(r, 5).value),
            'status1': clean(ws.cell(r, 6).value),
            'poc2': clean(ws.cell(r, 7).value),
            'status2': clean(ws.cell(r, 8).value),
        })
    wb2.close()

    unmatched_audit = []
    for key, entries in audit.items():
        if key not in funds:
            unmatched_audit.append(entries[0].get('poc1') or key)
            continue
        f = funds[key]
        for e in entries:
            for who, status_cell, is_first in ((e['poc1'], e['status1'], True), (e['poc2'], e['status2'], False)):
                if not who:
                    continue
                people = split_people(who)
                statuses = split_statuses(status_cell)
                for i, person in enumerate(people):
                    emp, note = status_to_employment(statuses[i] if i < len(statuses) else (statuses[0] if statuses else None))
                    match = next((c for c in f['contacts'] if norm_name(c['name']) == norm_name(person)), None)
                    if match is None:
                        match = {'name': person, 'role': None, 'linkedin_url': None, 'linkedin_status': None,
                                 'phone': None, 'email': None, 'slot': 1 if is_first else 2, 'multi_index': i}
                        f['contacts'].append(match)
                    match['employment_status'] = emp
                    match['audit_note'] = note
                    if is_first:
                        match['role'] = match['role'] or e['designation']
                        match['new_company'] = e['new_company']
                        match['new_designation'] = e['new_designation']

    # ── Normalise into the final shape ──
    out = []
    for key, f in funds.items():
        inc, exc, left = parse_sectors(f['sectors_raw'])
        smin, smax, stage_unparsed = parse_stage(f['stage_raw'])
        tmin, tmax, cur, ticket_unparsed = parse_ticket(f['ticket_raw'])

        st_raw = (f['service_type_raw'] or '').strip().lower().rstrip('s')
        service_type = SERVICE_TYPE_MAP.get(st_raw)
        if not service_type:
            for k, v in SERVICE_TYPE_MAP.items():
                if k in st_raw:
                    service_type = v
                    break
        if not service_type:
            service_type = 'vc_fund'
            if f['service_type_raw']:
                report['service_type_guessed'].append(f"{f['name']}: {f['service_type_raw']!r} -> vc_fund")

        notes = list(f['notes'])
        if ticket_unparsed:
            notes.append(f"Ticket size (currency not stated in source, needs review): {ticket_unparsed}")
            report['ticket_no_currency'].append(f"{f['name']}: {ticket_unparsed}")
        if stage_unparsed:
            notes.append(f"Stage (source text): {stage_unparsed}")
            report['stage_unparsed'].append(f"{f['name']}: {stage_unparsed}")
        if left:
            report['sector_unmapped'].extend(left)

        # Primary / secondary. Order of preference: audit slot, then whether they are still there —
        # a contact who has left should not be the primary POC anyone is told to call.
        contacts = f['contacts']
        for c in contacts:
            c.setdefault('employment_status', 'unknown')
            c.setdefault('new_company', None)
            c.setdefault('new_designation', None)
            c.setdefault('audit_note', None)
        contacts.sort(key=lambda c: (
            0 if c['employment_status'] == 'active' else 1 if c['employment_status'] == 'unknown' else 2,
            c.get('slot', 9), c.get('multi_index', 0),
        ))
        for i, c in enumerate(contacts):
            c['rank'] = 'primary' if i == 0 else ('secondary' if i == 1 else 'other')

        out.append({
            'name': f['name'],
            'sources': f['sources'],
            'country': f['country'],
            'website': f['website'],
            'service_type': service_type,
            'service_type_raw': f['service_type_raw'],
            'sectors': inc,
            'excluded_sectors': exc,
            'stage_min': smin, 'stage_max': smax, 'stage_raw': f['stage_raw'],
            'ticket_min': tmin, 'ticket_max': tmax, 'ticket_currency': cur, 'ticket_raw': f['ticket_raw'],
            'connect_strength': f['connect_strength'],
            'esv_pocs': f['esv_pocs'],
            'notes': notes,
            'contacts': contacts,
        })

    out.sort(key=lambda x: x['name'].lower())
    json.dump(out, open(os.path.join(OUT, 'funds.json'), 'w', encoding='utf-8'), indent=1, ensure_ascii=False)

    # ── Report ──
    L = []
    L.append(f"# Fund import — collation report\n")
    L.append(f"**{len(out)} unique funds** collated from 3 sheets + the employment audit.\n")
    src = Counter(tuple(sorted(f['sources'])) for f in out)
    L.append("## Where each fund came from\n")
    for k, v in src.most_common():
        L.append(f"- {v:>3}  {' + '.join(k)}")
    contacts_total = sum(len(f['contacts']) for f in out)
    emp = Counter(c['employment_status'] for f in out for c in f['contacts'])
    L.append(f"\n## POCs\n\n- {contacts_total} contacts across {len(out)} funds")
    for k, v in emp.most_common():
        L.append(f"- {v:>3}  {k}")
    L.append(f"- {sum(1 for f in out for c in f['contacts'] if c.get('new_company'))} have a known new employer")
    L.append(f"- {sum(1 for f in out if not f['contacts'])} funds have NO contact at all")

    poc = Counter(i for f in out for i in f['esv_pocs'])
    L.append("\n## ESV attribution\n")
    for k, v in poc.most_common():
        nm, cur_user = ESV_PEOPLE[k]
        L.append(f"- {v:>3}  {k} — {nm}{'' if cur_user else '  *(left; kept as a name, no user link)*'}")
    L.append(f"- {sum(1 for f in out if not f['esv_pocs'])} funds have no ESV POC recorded")

    L.append("\n## Coverage\n")
    for field, pred in [
        ('website', lambda f: f['website']), ('country', lambda f: f['country']),
        ('sectors', lambda f: f['sectors']), ('excluded sectors', lambda f: f['excluded_sectors']),
        ('stage range', lambda f: f['stage_min']), ('ticket size', lambda f: f['ticket_min']),
        ('connect strength', lambda f: f['connect_strength']),
    ]:
        n = sum(1 for f in out if pred(f))
        L.append(f"- {n:>3}/{len(out)}  {field}")

    L.append("\n## Needs a human\n")
    L.append(f"\n### Ticket size with no currency in the source ({len(report['ticket_no_currency'])})")
    L.append("Imported as blank, original text kept in the fund's notes.\n")
    for x in report['ticket_no_currency'][:40]:
        L.append(f"- {x}")
    if len(report['ticket_no_currency']) > 40:
        L.append(f"- …and {len(report['ticket_no_currency'])-40} more")

    L.append(f"\n### Stage text that did not parse ({len(report['stage_unparsed'])})")
    for x in report['stage_unparsed'][:25]:
        L.append(f"- {x}")

    L.append(f"\n### Service type guessed as VC Fund ({len(report['service_type_guessed'])})")
    for x in report['service_type_guessed'][:25]:
        L.append(f"- {x}")

    um = Counter(report['sector_unmapped'])
    L.append(f"\n### Sector words with no canonical tag ({len(um)} distinct)")
    L.append("Left off the tags rather than invented. Add any that matter to CANON and re-run.\n")
    for w, n in um.most_common(40):
        L.append(f"- {n:>3}  {w}")

    if unmatched_audit:
        L.append(f"\n### Audit rows whose fund was not found ({len(unmatched_audit)})")
        for x in unmatched_audit:
            L.append(f"- {x}")

    open(os.path.join(OUT, 'report.md'), 'w', encoding='utf-8').write('\n'.join(L) + '\n')
    print(f"{len(out)} funds -> funds.json")
    print(f"{contacts_total} contacts, {emp.get('active',0)} active / {emp.get('moved_on',0)} moved on")
    print(f"report -> {os.path.join(OUT,'report.md')}")


if __name__ == '__main__':
    main()
